"""Read official Hacienda/Finanzas workbooks; never modify originals.
Usage: python scripts_import_national_operations.py imig.xlsx deuda.xlsx
Requires openpyxl. Fixed July 2026 publication: fail if layout/period changes.
"""
import csv,json,sys,hashlib
from pathlib import Path
ROOT=Path(__file__).resolve().parent
FISCAL_URL='https://www.argentina.gob.ar/sites/default/files/2026/08/imig_julio_2026.xlsx'
DEBT_URL='https://www.argentina.gob.ar/sites/default/files/deuda_publica_31-03-2026.xlsx'
def build(fiscal_path,debt_path):
    import openpyxl
    fiscal=openpyxl.load_workbook(fiscal_path,data_only=True);sheet=fiscal['Julio'];monthly=fiscal['Mensualizacion']
    with (ROOT/'data/ipc_national_index.csv').open() as f:ipc={r['period']:float(r['ipc_index']) for r in csv.DictReader(f)}
    assert any(str(c.value).startswith('2026-07-01') for row in sheet for c in row)
    groups={'Prestaciones sociales':'social','Subsidios económicos':'subsidies','Gastos de funcionamiento y otros':'operating','Transferencias corrientes a provincias':'provinces','Transferencias a universidades':'universities','Otros Gastos Corrientes':'other','Gastos de capital':'capital'}
    selected={**groups,'INGRESOS TOTALES':'income','GASTOS PRIMARIOS':'primary_spend','RESULTADO PRIMARIO':'primary','Intereses Netos':'interest','RESULTADO FINANCIERO':'financial','Jubilaciones y pensiones contributivas':'pensions','Asignación Universal para Protección Social':'auh','Salarios':'salaries'}
    rows=[];seen=set()
    for cells in sheet:
        label=next((str(c.value).strip() for c in cells[1:4] if c.value is not None),'')
        if label not in selected or label in seen:continue
        seen.add(label);value=cells[6].value;previous=cells[7].value
        assert isinstance(value,(float,int)) and isinstance(previous,(float,int))
        rows.append(dict(id=selected[label],label=label,value=value,previous=previous,nominal_yoy_pct=100*(value/previous-1) if previous else None,real_yoy_pct=100*((value/previous)/(ipc['2026-07']/ipc['2025-07'])-1) if previous else None,ytd=cells[11].value,composition=label in groups,source_sheet=sheet.title,source_range=f'G{cells[6].row}:H{cells[6].row}'))
    assert len(rows)==len(selected),(len(rows),len(selected))
    indexed={r['id']:r for r in rows};total=indexed['primary_spend']['value']
    assert abs(sum(r['value'] for r in rows if r['composition'])-total)<1
    assert abs(indexed['income']['value']-total-indexed['primary']['value'])<1
    assert abs(indexed['primary']['value']-indexed['interest']['value']-indexed['financial']['value'])<1
    history=[]
    for cells in monthly:
        label=next((str(c.value).strip() for c in cells[1:4] if c.value is not None),'')
        if label not in selected:continue
        for col in range(6,13):
            value=cells[col].value;assert isinstance(value,(float,int))
            period=f'2026-{col-5:02}'
            history.append(dict(id=selected[label],period=period,value=value,real_july2026=value*ipc['2026-07']/ipc[period],source_cell=f'{cells[col].column_letter}{cells[col].row}'))
    assert len(history)==len(rows)*7
    debt=openpyxl.load_workbook(debt_path,data_only=True)['A.3.1']
    assert 'miles' in debt['B10'].value and '31/03/2026' in debt['B10'].value
    maturities=[]
    for col,period in list(zip(range(3,12),[f'2026-{m:02}' for m in range(4,13)]))+list(zip(range(13,16),[f'2027-{m:02}' for m in range(1,4)])):
        values={key:debt.cell(row,col).value/1000 for key,row in [('total',59),('capital',60),('interest',61)]}
        assert abs(values['total']-values['capital']-values['interest'])<.00001
        maturities.append(dict(period=period,**values,source_range=f'{debt.cell(59,col).coordinate}:{debt.cell(61,col).coordinate}'))
    return {'reviewed_at':'2026-09-06','spending':{'period':'2026-07','publication_date':'2026-08-18','scope':'Sector Público Nacional, base caja','unit':'ARS millones','source_url':FISCAL_URL,'sha256':hashlib.sha256(Path(fiscal_path).read_bytes()).hexdigest(),'rows':rows,'monthly_2026':history,'real_method':'Variación julio/julio deflactada con IPC nacional. Serie mensual 2026 a precios de julio 2026; no desestacionalizada.','transfer_scope':'La fila de transferencias corrientes a provincias del IMIG corresponde a Administración Nacional. No equivale al total SPN ni incluye coparticipación. Los componentes se muestran según la clasificación del archivo.'},'maturities':{'as_of':'2026-03-31','scope':'Deuda bruta de la Administración Central en situación de pago normal','unit':'USD millones equivalentes','source_url':DEBT_URL,'source_sheet':'A.3.1','sha256':hashlib.sha256(Path(debt_path).read_bytes()).hexdigest(),'includes_bcra_advances':True,'reading':'Perfil estático del stock al 31/03/2026, con tipos de cambio de esa fecha. Incluye adelantos BCRA. No incorpora nuevas emisiones, canjes ni pagos posteriores. Excluye pagos eventuales vinculados al PIB. No constituye un calendario actualizado al día de hoy.','rows':maturities}}
if __name__=='__main__':
    result=build(sys.argv[1],sys.argv[2]);(ROOT/'data/national_operations.json').write_text(json.dumps(result,ensure_ascii=False,indent=2),encoding='utf-8');print('Validated spending components and 12 maturity months')
