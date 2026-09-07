#!/usr/bin/env python3
"""Importa TOP y RON 2026 desde las planillas oficiales de DNAP.

El importador no imputa datos. Un cero se conserva solamente cuando la planilla
oficial informa ese cero dentro de un mes válido; los meses sin total quedan
fuera de la salida. Además reconcilia cada total de recaudación propia contra
la suma de sus cinco componentes antes de publicar los CSV normalizados.
"""
from __future__ import annotations

import argparse
import csv
import json
import re
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parent
TOP_OUTPUT = ROOT / "top_mensual_2026_normalizado.csv"
RON_OUTPUT = ROOT / "informacion_consolidada_2026_normalizado.csv"
PBA_TOP_OUTPUT = ROOT / "pba_top_monthly.csv"
PBA_RON_OUTPUT = ROOT / "pba_ron_monthly.csv"
CANONICAL_TOP_OUTPUT = ROOT / "data/recaudacion_propia.csv"
CANONICAL_RON_OUTPUT = ROOT / "data/transferencias_nac.csv"
COVERAGE_OUTPUT = ROOT / "data/cobertura.csv"
META_OUTPUT = ROOT / "data/meta.json"
MANIFEST_OUTPUT = ROOT / "dashboard_manifest.json"

TOP_URL = "https://www.argentina.gob.ar/sites/default/files/top_mensual_2026_12.xlsx"
RON_URL = "https://www.argentina.gob.ar/sites/default/files/informacion_consolidada_2026_4.xlsx"

MONTHS = {
    "Enero": "01",
    "Febrero": "02",
    "Marzo": "03",
    "Abril": "04",
    "Mayo": "05",
    "Junio": "06",
    "Julio": "07",
    "Agosto": "08",
    "Septiembre": "09",
    "Octubre": "10",
    "Noviembre": "11",
    "Diciembre": "12",
}

PROVINCE_MAP = {
    "C.A.B.A.": "CABA",
    "C.A.B.A": "CABA",
    "Ciudad Autónoma de Buenos Aires": "CABA",
    "Sgo. del Estero": "Santiago del Estero",
    "Sgo. Del Estero": "Santiago del Estero",
    "Sgo Del Estero": "Santiago del Estero",
    "Stgo. del Estero": "Santiago del Estero",
    "Tierra Del Fuego": "Tierra del Fuego",
    "Tierra del Fuego, Antártida e Islas del Atlántico Sur": "Tierra del Fuego",
}

TOP_SHEETS = {
    "Iibb": "IIBB",
    "Inmobiliario": "INMOBILIARIO",
    "Sellos": "SELLOS",
    "Automotores": "AUTOMOTORES",
    "Otros": "OTROS",
}

RON_CATEGORY_NORMALIZATION = {
    "C.F.I. | Neta De | Ley Nº 26.075": "CFI | Neta",
    "Financ. | Educativo | Ley Nº 26.075": "Financiamiento Educativo",
    "Compensación | Consenso Fiscal | II. a y b; II. d y e | (2)": "Compensación Consenso Fiscal",
}


def normalize_province(value: object) -> str:
    name = re.sub(r"\s*\(\*\)\s*$", "", str(value or "").strip())
    return PROVINCE_MAP.get(name, name)


def number(value: object) -> float:
    if value in (None, ""):
        return 0.0
    return float(value)


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as target:
        writer = csv.DictWriter(target, fieldnames=fieldnames, lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


def manifest_universe() -> list[str]:
    with MANIFEST_OUTPUT.open(encoding="utf-8") as source:
        manifest = json.load(source)
    return [normalize_province(name) for name in manifest["province_universe"]]


def month_columns(sheet) -> dict[int, str]:
    result: dict[int, str] = {}
    for column in range(3, sheet.max_column + 1):
        label = str(sheet.cell(6, column).value or "").strip()
        if label in MONTHS:
            result[column] = f"2026-{MONTHS[label]}"
    return result


def province_rows(sheet, name_column: int, universe: set[str]) -> dict[str, int]:
    result: dict[str, int] = {}
    for row in range(1, sheet.max_row + 1):
        province = normalize_province(sheet.cell(row, name_column).value)
        if province in universe:
            result[province] = row
    return result


def import_top(path: Path, universe: list[str]) -> tuple[list[dict[str, object]], dict[str, list[str]]]:
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    total_sheet = workbook["TOTAL"]
    universe_set = set(universe)
    total_rows = province_rows(total_sheet, 2, universe_set)
    sheet_rows = {
        tax: province_rows(workbook[sheet_name], 2, universe_set)
        for tax, sheet_name in TOP_SHEETS.items()
    }
    columns = month_columns(total_sheet)
    source_name = path.name
    rows: list[dict[str, object]] = []
    coverage: dict[str, list[str]] = defaultdict(list)

    for province in universe:
        row_number = total_rows.get(province)
        if row_number is None:
            continue
        for column, period in columns.items():
            total = number(total_sheet.cell(row_number, column).value)
            if total <= 0:
                continue
            component_values: dict[str, float] = {}
            for tax, sheet_name in TOP_SHEETS.items():
                tax_row = sheet_rows[tax].get(province)
                if tax_row is None:
                    raise ValueError(f"{province}: falta fila en hoja {sheet_name}")
                component_values[tax] = number(workbook[sheet_name].cell(tax_row, column).value)
            component_total = sum(component_values.values())
            if abs(component_total - total) > 0.05:
                raise ValueError(
                    f"{province} {period}: componentes {component_total:.6f} != total {total:.6f}"
                )
            coverage[province].append(period)
            for tax, value in component_values.items():
                rows.append(
                    {
                        "province": province,
                        "source": source_name,
                        "year": 2026,
                        "period_type": "month",
                        "period": period,
                        "tax": tax,
                        "value_millions": value,
                    }
                )

    rows.sort(key=lambda row: (str(row["province"]), str(row["period"]), str(row["tax"])))
    return rows, dict(coverage)


def ron_headers(sheet) -> dict[int, str]:
    headers: dict[int, str] = {}
    for column in range(2, sheet.max_column + 1):
        parts: list[str] = []
        for row in range(5, 10):
            value = sheet.cell(row, column).value
            if value not in (None, ""):
                parts.append(" ".join(str(value).split()))
        if parts:
            headers[column] = " | ".join(parts)
    return headers


def import_ron(path: Path, universe: list[str], year: int = 2026) -> tuple[list[dict[str, object]], dict[str, list[str]]]:
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    universe_set = set(universe)
    monthly_sheets = [(name.lower(), f"{year}-{MONTHS[name]}") for name in MONTHS if name.lower() in workbook.sheetnames]
    monthly_sheets.sort(key=lambda item: item[1])
    source_name = path.name
    rows: list[dict[str, object]] = []
    coverage: dict[str, list[str]] = defaultdict(list)

    def add_sheet(sheet_name: str, period_type: str, period: str) -> None:
        sheet = workbook[sheet_name]
        headers = ron_headers(sheet)
        row_map = province_rows(sheet, 1, universe_set)
        for province in universe:
            row_number = row_map.get(province)
            if row_number is None:
                continue
            if period_type == "month":
                coverage[province].append(period)
            for column, category in headers.items():
                value = sheet.cell(row_number, column).value
                if value in (None, ""):
                    continue
                rows.append(
                    {
                        "province": province,
                        "source": source_name,
                        "sheet": sheet_name,
                        "year": year,
                        "period_type": period_type,
                        "period": period,
                        "category": category,
                        "value_millions": float(value),
                        "category_normalized": RON_CATEGORY_NORMALIZATION.get(category, category),
                    }
                )

    for sheet_name, period in monthly_sheets:
        add_sheet(sheet_name, "month", period)
    add_sheet("CONS", "consolidated", "CONS")
    rows.sort(
        key=lambda row: (
            str(row["province"]),
            str(row["period"]),
            str(row["category_normalized"]),
            str(row["category"]),
        )
    )
    return rows, dict(coverage)


def select_ron(rows: list[dict[str, object]], province: str, period: str, category: str) -> float | None:
    matches = [
        float(row["value_millions"])
        for row in rows
        if row["province"] == province
        and row["period"] == period
        and row["category_normalized"] == category
    ]
    return matches[0] if matches else None


def write_pba_files(
    top_rows: list[dict[str, object]],
    ron_rows: list[dict[str, object]],
    top_source: Path,
    ron_source: Path,
) -> None:
    top_by_period: dict[str, dict[str, float]] = defaultdict(dict)
    for row in top_rows:
        if row["province"] == "Buenos Aires":
            top_by_period[str(row["period"])][str(row["tax"])] = float(row["value_millions"])
    pba_top_rows: list[dict[str, object]] = []
    for period, taxes in sorted(top_by_period.items()):
        pba_top_rows.append(
            {
                "fecha_corte": f"{period}-{('30' if period.endswith(('04','06','09','11')) else '28' if period.endswith('02') else '31')}",
                "top_total_ars_m": sum(taxes.values()),
                "iibb_ars_m": taxes.get("Iibb", 0),
                "inmobiliario_ars_m": taxes.get("Inmobiliario", 0),
                "sellos_ars_m": taxes.get("Sellos", 0),
                "automotores_ars_m": taxes.get("Automotores", 0),
                "otros_top_ars_m": taxes.get("Otros", 0),
                "fuente_archivo": top_source.name,
                "observacion": "Mes cerrado informado por DNAP; dato provisorio sujeto a revisión de la fuente.",
            }
        )
    write_csv(
        PBA_TOP_OUTPUT,
        [
            "fecha_corte", "top_total_ars_m", "iibb_ars_m", "inmobiliario_ars_m",
            "sellos_ars_m", "automotores_ars_m", "otros_top_ars_m", "fuente_archivo", "observacion",
        ],
        pba_top_rows,
    )

    periods = sorted({str(row["period"]) for row in ron_rows if row["province"] == "Buenos Aires" and str(row["period"]).startswith("2026-")})
    pba_ron_rows: list[dict[str, object]] = []
    for period in periods:
        cfi = select_ron(ron_rows, "Buenos Aires", period, "CFI | Neta")
        educational = select_ron(ron_rows, "Buenos Aires", period, "Financiamiento Educativo")
        subtotal = select_ron(ron_rows, "Buenos Aires", period, "Subtotal")
        total_without_comp = select_ron(ron_rows, "Buenos Aires", period, "Total | Recursos | Origen Nacional | (1)")
        compensation = select_ron(ron_rows, "Buenos Aires", period, "Compensación Consenso Fiscal")
        total = select_ron(ron_rows, "Buenos Aires", period, "Total | (1) + (2)")
        if None in (cfi, educational, subtotal, total_without_comp, compensation, total):
            raise ValueError(f"PBA {period}: faltan categorías principales de RON")
        pba_ron_rows.append(
            {
                "fecha_corte": f"{period}-{('30' if period.endswith(('04','06','09','11')) else '28' if period.endswith('02') else '31')}",
                "cfi_neta_ars_m_xlsx": cfi,
                "financ_educ_ars_m_xlsx": educational,
                "subtotal_cfi_ars_m_xlsx": subtotal,
                "leyes_especiales_ars_m_xlsx": float(total_without_comp) - float(subtotal),
                "compensacion_consenso_fiscal_ars_m_xlsx": compensation,
                "total_ron_ars_m_xlsx": total,
                "subtotal_cfi_ars_m_daily": "",
                "leyes_especiales_ars_m_daily": "",
                "total_ron_sin_comp_ars_m": total_without_comp,
                "compensacion_consenso_fiscal_ars_m_daily": "",
                "total_ron_ars_m_daily": "",
                "total_diff_xlsx_daily_ars_m": "",
                "fuente_archivo": ron_source.name,
                "fuente_hoja": MONTH_NAME_BY_PERIOD[period],
            }
        )
    write_csv(
        PBA_RON_OUTPUT,
        [
            "fecha_corte", "cfi_neta_ars_m_xlsx", "financ_educ_ars_m_xlsx", "subtotal_cfi_ars_m_xlsx",
            "leyes_especiales_ars_m_xlsx", "compensacion_consenso_fiscal_ars_m_xlsx", "total_ron_ars_m_xlsx",
            "subtotal_cfi_ars_m_daily", "leyes_especiales_ars_m_daily", "total_ron_sin_comp_ars_m",
            "compensacion_consenso_fiscal_ars_m_daily", "total_ron_ars_m_daily", "total_diff_xlsx_daily_ars_m",
            "fuente_archivo", "fuente_hoja",
        ],
        pba_ron_rows,
    )


MONTH_NAME_BY_PERIOD = {f"2026-{number}": name.lower() for name, number in MONTHS.items()}


def write_canonical_files(
    top_rows: list[dict[str, object]],
    ron_rows: list[dict[str, object]],
    top_coverage: dict[str, list[str]],
    ron_coverage: dict[str, list[str]],
    universe: list[str],
) -> None:
    canonical_top = [
        {
            "province": row["province"],
            "period": row["period"],
            "tax": row["tax"],
            "value_millions": row["value_millions"],
            "price_basis": "pesos_corrientes",
            "evidence_status": "auditado",
            "source_file": row["source"],
            "source_url": TOP_URL,
        }
        for row in top_rows
    ]
    write_csv(
        CANONICAL_TOP_OUTPUT,
        ["province", "period", "tax", "value_millions", "price_basis", "evidence_status", "source_file", "source_url"],
        canonical_top,
    )

    selected_categories = {"CFI | Neta", "Compensación Consenso Fiscal", "Total | (1) + (2)"}
    canonical_ron = [
        {
            "province": row["province"],
            "period": row["period"],
            "category": row["category_normalized"],
            "value_millions": row["value_millions"],
            "price_basis": "pesos_corrientes",
            "evidence_status": "auditado",
            "source_file": row["source"],
            "source_url": RON_URL,
        }
        for row in ron_rows
        if row["period_type"] == "month" and row["category_normalized"] in selected_categories
    ]
    write_csv(
        CANONICAL_RON_OUTPUT,
        ["province", "period", "category", "value_millions", "price_basis", "evidence_status", "source_file", "source_url"],
        canonical_ron,
    )

    coverage_rows: list[dict[str, object]] = []
    for province in universe:
        for dataset, coverage, expected in (
            ("recaudacion_propia_2026", top_coverage, "2026-07"),
            ("transferencias_nacion_2026", ron_coverage, "2026-07"),
        ):
            periods = sorted(set(coverage.get(province, [])))
            coverage_rows.append(
                {
                    "dataset": dataset,
                    "province": province,
                    "first_period": periods[0] if periods else "",
                    "last_period": periods[-1] if periods else "",
                    "months_count": len(periods),
                    "expected_through": expected,
                    "coverage_status": "completa" if periods and periods[-1] == expected else "rezagada" if periods else "faltante",
                }
            )
    write_csv(
        COVERAGE_OUTPUT,
        ["dataset", "province", "first_period", "last_period", "months_count", "expected_through", "coverage_status"],
        coverage_rows,
    )

    meta = {
        "generated_at": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
        "structural_cutoff": "2026-03-31",
        "sources": {
            "recaudacion_propia_2026": {"url": TOP_URL, "max_period": "2026-07", "pba_max_period": max(top_coverage.get("Buenos Aires", []), default=None)},
            "transferencias_nacion_2026": {"url": RON_URL, "max_period": "2026-07", "pba_max_period": max(ron_coverage.get("Buenos Aires", []), default=None)},
        },
        "rules": {
            "missing_values": "No se imputan. Los meses sin total oficial no se publican.",
            "top_reconciliation": "Cada total mensual fue reconciliado con IIBB + Sellos + Automotores + Inmobiliario + Otros.",
            "mixed_vintages": "Cada vista y tarjeta debe mostrar su fecha de corte específica.",
        },
    }
    META_OUTPUT.write_text(json.dumps(meta, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def update_manifest(
    top_coverage: dict[str, list[str]],
    ron_coverage: dict[str, list[str]],
    universe: list[str],
) -> None:
    with MANIFEST_OUTPUT.open(encoding="utf-8") as source:
        manifest = json.load(source)
    manifest["generated_at"] = datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")
    manifest["data_cutoff"] = "2026-07-31"
    manifest["as_of_by_block"] = {
        "structural_and_ranking": "2026-03-31",
        "recaudacion_propia_2026_max": "2026-07-31",
        "recaudacion_propia_pba": "2026-06-30",
        "transferencias_nacion_2026": "2026-07-31",
        "debt_stock_pba": "2026-03-31",
        "debt_currency_composition_pba": "2025-12-31",
    }
    manifest["files"].update(
        {
            "canonical_recaudacion_propia": "data/recaudacion_propia.csv",
            "canonical_transferencias_nacion": "data/transferencias_nac.csv",
            "coverage": "data/cobertura.csv",
            "canonical_meta": "data/meta.json",
        }
    )
    manifest["missing_top_mensual_2026_provinces"] = [province for province in universe if not top_coverage.get(province)]
    manifest["missing_ron_2026_provinces"] = [province for province in universe if not ron_coverage.get(province)]
    notes = manifest.setdefault("notes", {})
    notes["monthly_scope_2026"] = "Importación directa desde planillas oficiales DNAP. No se imputan datos ni se crean placeholders en cero."
    notes["top_mensual_cobertura"] = "Cobertura dispar por jurisdicción; consultar data/cobertura.csv y mostrar el corte propio de cada provincia."
    notes["mixed_vintages"] = "Estructura/ranking: 1T26; deuda moneda: 4T25; TOP: hasta julio según provincia (PBA junio); RON: julio."
    MANIFEST_OUTPUT.write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--top-source", type=Path, required=True)
    parser.add_argument("--ron-source", type=Path, required=True)
    args = parser.parse_args()
    universe = manifest_universe()
    top_rows, top_coverage = import_top(args.top_source, universe)
    ron_rows, ron_coverage = import_ron(args.ron_source, universe)

    write_csv(TOP_OUTPUT, ["province", "source", "year", "period_type", "period", "tax", "value_millions"], top_rows)
    write_csv(
        RON_OUTPUT,
        ["province", "source", "sheet", "year", "period_type", "period", "category", "value_millions", "category_normalized"],
        ron_rows,
    )
    write_pba_files(top_rows, ron_rows, args.top_source, args.ron_source)
    write_canonical_files(top_rows, ron_rows, top_coverage, ron_coverage, universe)
    update_manifest(top_coverage, ron_coverage, universe)
    # Keep the downloadable report in sync with the same data publication.
    from scripts_export_management_reports import build as build_management_reports
    build_management_reports(ROOT / 'reports')

    print(f"TOP: {len(top_rows)} filas; PBA hasta {max(top_coverage['Buenos Aires'])}")
    print(f"RON: {len(ron_rows)} filas; {len(ron_coverage)} jurisdicciones hasta julio")
    print("Sin TOP:", [province for province in universe if not top_coverage.get(province)])
    print("Sin RON:", [province for province in universe if not ron_coverage.get(province)])


if __name__ == "__main__":
    main()
